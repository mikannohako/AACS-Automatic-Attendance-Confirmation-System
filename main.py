
#? 各ライブラリのインポート
import PySimpleGUI as sg
import sys
import os
from datetime import datetime
from tkinter import messagebox
from tkinter import simpledialog
import tkinter as tk
import tkinter.simpledialog as simpledialog
import openpyxl
from openpyxl import load_workbook
from openpyxl.worksheet.table import TableStyleInfo
from openpyxl.worksheet.table import Table
import json
import logging
import requests
import webbrowser
import msvcrt
import tempfile
import cv2
import face_recognition
import numpy as np
import pickle  # データ保存用

#? tkinterのダイアログボックスを常時最前面に表示

root = tk.Tk()
root.attributes('-topmost', True)
root.withdraw()

#? 重複起動禁止関係

'''
一時ディレクトリにロックファイルを作成して
ロックファイルが存在してたら起動不可、存在してなかったら起動可能。
問題点は異常終了した場合にロックファイルが消されなくて次回起動できなくなるのと
終了時に毎回削除の関数を置かなくちゃいけなくなること。
'''

# 一時ディレクトリにロックファイルを作成
lock_file_path = os.path.join(tempfile.gettempdir(), 'main.lock')

try:
    # ロックファイルを開く
    lock_file = open(lock_file_path, 'w')
    
    # ロックを取得する
    msvcrt.locking(lock_file.fileno(), msvcrt.LK_NBLCK, 1)
except IOError:
    # ロック取得に失敗した場合は他のインスタンスが実行中
    messagebox.showerror("Error", "複数同時起動はできません。")

    # エラーで終了
    sys.exit(1)

# 起動用のプログレスバーの最大値を代入
BAR_MAX = 100

# プログレスバーのUIを作成
layout = [
    [sg.Text('起動中')],
    [sg.ProgressBar(BAR_MAX, orientation='h', size=(20, 20), key='-PROG-')],
    ]

# 作ったUIを表示
window = sg.Window('起動中', layout, keep_on_top=True)

# ウィンドウを初期化
event, values = window.read(timeout=0)
if event == sg.WINDOW_CLOSED:
    window.close()

# プログレスバーの値を10に設定
window['-PROG-'].update(10)

#? ログの設定

# ログファイルの出力パス
filename = 'logfile.log'

# ログのメッセージフォーマットを指定
fmt = "%(asctime)s - %(levelname)s - %(message)s - %(module)s - %(funcName)s - %(lineno)d"

# ログの出力レベルを設定
logging.basicConfig(filename=filename, encoding='utf-8', level=logging.INFO, format=fmt)

# プログレスバーの値を20に設定
window['-PROG-'].update(20)

#? 各機能の関数

def exit_with_error(message): #? エラー時の処理用関数
    # ログのメッセージを作成
    logging.critical(f"{message}")
    # エラーダイアログボックスを表示
    messagebox.showerror("Error", f"エラーが発生しました。\n{message}")
    
    #? 重複起動関係
        
    # ロックを解放する
    msvcrt.locking(lock_file.fileno(), msvcrt.LK_UNLCK, 1)
    # ロックファイルを閉じる
    lock_file.close()
    # ロックファイルを削除する
    os.remove(lock_file_path)
    
    sys.exit(1)

# 外部Excelファイルを読み込む
def load_user_data(file_path):
    # user_list.xlsxファイルを開く
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook["user_list"]  # user_listシートを指定
    
    # データを読み込む
    all_data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):  # 1行目はヘッダーなので2行目から読み込む
        user_id, user_name = row
        all_data.append((user_name, user_id))  # 名前とIDをタプルで追加
    
    return all_data

def update(): #? アップデート

    def check_internet_connection(): # ネットにつながっているか確認
        try:
            # Googleに接続
            response = requests.get("http://www.google.com", timeout=5)
            # HTTPエラーコードが返ってきた場合に例外を発生させる。
            response.raise_for_status()
            
            return True
        except requests.RequestException as e: # 接続時に例外が発生した場合の処理
            logging.warning("インターネット接続の確立に失敗しました。: %s", e)
            return False
    
    def update_check(): # 最新バージョンがリリースされているかを確認
        # 接続用のURLを代入
        api_url = f"https://api.github.com/repos/mikannohako/AACS-Automatic-Attendance-Confirmation-System/releases/latest"
        
        # 最新のリリース情報を取得
        response = requests.get(api_url)
        # HTTPリクエストのレスポンスステータスコードが200（成功）になっているかを確認
        if response.status_code == 200:
            # 帰ってきた情報をjson形式でパースして変数に代入
            release_info = response.json()
            
            # パースされたjsonデータからバージョン情報を取得して変数に格納
            tag_name = release_info["tag_name"]
            
            # 不要な文字を取り除いて整数にして代入
            tag_name_int = int(tag_name.replace("v", "").replace(".", ""))
            
            if config_data["version"] < tag_name_int: # コンフィグデータから現在バージョンを取得して最新バージョンより小さいかを確認
                if messagebox.askyesno("更新", "新しいバージョンがリリースされています。\n更新してください。"): # 更新するかを確認
                    
                    # 最新のバージョンのリリースブラウザで開く
                    url = 'https://github.com/mikannohako/AACS-Automatic-Attendance-Confirmation-System/releases/latest'
                    webbrowser.open(url)
                    
                    #? 終了
                    
                    # ロックを解放する
                    msvcrt.locking(lock_file.fileno(), msvcrt.LK_UNLCK, 1)
                    # ロックファイルを閉じる
                    lock_file.close()
                    # ロックファイルを削除する
                    os.remove(lock_file_path)
                    
                    sys.exit(0)
    
    if check_internet_connection(): # ネット接続を確認したらupdate_check関数を実行
        update_check()

def json_save(): #? JSONデータを保存
    # jsonファイルをindent=4で保存
    with open('config.json', 'w') as f:
        json.dump(config_data, f, indent=4)

# ログ設定
logging.basicConfig(level=logging.INFO)

def record_file_creation():  #? 記録ファイル作成
    def create_month_sheet(workbook, month):
        sheet_name = month
        if month.startswith("0"):
            sheet_name = month[1:]
        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(sheet_name)
            return workbook[sheet_name]

    def input_date_data(sheet, month):
        days_in_month = 30 if month in ["04", "06", "09", "11"] else 31 if month != "02" else 29
        for day in range(1, days_in_month + 1):
            sheet.cell(row=1, column=day + 9).value = day
        headers = ['名前', 'ID', '出席率', '全日数', '出席', '欠席', '無断欠席', '遅刻', '早退']
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col).value = header

    # 現在の年を取得
    current_date_y = datetime.now().year
    
    # ユーザーリストをロード
    user_list_wb = openpyxl.load_workbook('user_list.xlsx')
    user_list_sheet = user_list_wb['user_list']

    # ユーザー情報を取得
    all_data = [
        (row[1], row[0])  # 名前, ID
        for row in user_list_sheet.iter_rows(min_row=2, max_col=2, values_only=True)
        if row[0] is not None and row[1] is not None  # 空の行を除外
    ]

    # 新しいWorkbook（エクセルファイル）を作成
    workbook = openpyxl.Workbook()

    for month in range(1, 13):
        month_str = str(month).zfill(2)
        sheet = create_month_sheet(workbook, f"{month_str}月")
        input_date_data(sheet, month_str)

        for data in all_data:
            row_number = sheet.max_row + 1
            sheet.cell(row=row_number, column=1, value=data[0])  # 名前
            sheet.cell(row=row_number, column=2, value=data[1])  # ID
            sheet.cell(row=row_number, column=3).value = f'=IFERROR((E{row_number} / D{row_number}) * 100, "No data")'
            sheet.cell(row=row_number, column=4).value = f'=COUNTIF(J{row_number}:BA{row_number}, "<>")'  # 全日数
            sheet.cell(row=row_number, column=5).value = f'=(COUNTIF(J{row_number}:BA{row_number}, "*出席*") + H{row_number} + I{row_number})'
            sheet.cell(row=row_number, column=6).value = f'=COUNTIF(J{row_number}:BA{row_number}, "*欠席*")'
            sheet.cell(row=row_number, column=7).value = f'=COUNTIF(J{row_number}:BA{row_number}, "無断欠席")'
            sheet.cell(row=row_number, column=8).value = f'=COUNTIF(J{row_number}:BA{row_number}, "*遅刻*")'
            sheet.cell(row=row_number, column=9).value = f'=COUNTIF(J{row_number}:BA{row_number}, "*早退*")'
        
        end_row = len(all_data) + 1
        table_range = f"A1:I{end_row}"
        table = Table(displayName=f"Table{month}", ref=table_range)
        
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showRowStripes=True
        )
        table.tableStyleInfo = style
        sheet.add_table(table)

    if "Sheet" in workbook.sheetnames:
        del workbook["Sheet"]

    workbook.save(f"{current_date_y}Attendance_records.xlsx")
    logging.info("記録用ファイルが作成されました。")

def record(): #? 出席
    #? config設定
    
    # 時間変数の設定
    current_date = datetime.now()
    current_date_y = current_date.strftime("%Y")
    current_date_d = current_date.strftime("%d")
    
    # 設定ファイルのパス
    config_file_path = 'config.json'
    
    # 設定ファイルの読み込み
    with open(config_file_path, 'r') as config_file:
        config_data = json.load(config_file)
    
    #? 変数の初期設定
    
    name = None
    
    current_date = datetime.now()
    
    #? 遅刻時間の設定
    while True:
        
        if config_data["automatic_late_time_setting"]:
            # 時間の設定が自動になっている場合
            
            # 各変数に現在の時間を代入
            lateness_time_hour = current_date.hour
            lateness_time_minute = current_date.minute
            # 設定されている時間分分を足す
            lateness_time_minute = lateness_time_minute + config_data['lateness_time']
            
            # 分が60以上だったら時間を一足して分から60引く
            if lateness_time_minute >= 60:
                lateness_time_minute = lateness_time_minute - 60
                lateness_time_hour = lateness_time_hour + 1
            
            # 確認メッセージボックス
            messagebox.showinfo("INFO", f"{lateness_time_hour}時{lateness_time_minute}分以降を遅刻として設定しました。")
            break
        else:
            # 時間の設定が手動になっている場合
            
            # 時間の設定の入力を求める
            lateness_time = simpledialog.askstring('入力してください。', '何時以降を遅刻と設定しますか？\n（HH:MMの形式で入力してください）')
            
            if lateness_time == None:
                return
            
            # 入力された時間を「:」で分割する
            try:
                hours, minutes = lateness_time.split(':')
            except ValueError:
                messagebox.showwarning("警告", "入力された値が正しい形式ではありません。")
                continue
            
            # 入力された時間を整数に変換する
            try:
                lateness_time_hour = int(hours)
                lateness_time_minute = int(minutes)
            except ValueError:
                messagebox.showwarning("警告", "入力された値が整数ではありません。整数値を入力してください。")
                continue
            
            # 入力された時間の確認
            if lateness_time_hour < 0 or lateness_time_hour > 23 or lateness_time_minute < 0 or lateness_time_minute > 59:
                messagebox.showwarning("警告", "無効な時間です。正しい形式で再度入力してください。")
                continue
            
            if messagebox.askokcancel("確認", f'{lateness_time_hour}時{lateness_time_minute}分から遅刻に設定しますか？'):
                messagebox.showinfo("INFO", f"{lateness_time_hour}時{lateness_time_minute}分以降を遅刻として設定しました。")
                break
            else:
                return
    
    #? Excel初期設定
    
    # 記録ファイル名
    ar_filename = f"{current_date_y}Attendance_records.xlsx"
    
    # 新しいWorkbook（エクセルファイル）を作成して、ファイル名を指定
    try:
        workbook = load_workbook(ar_filename)
    except FileNotFoundError:
        messagebox.showerror("Error: File not found", "記録用のファイルがありません。\nもう一度起動を行ってください。")
        
        #? 重複起動関係
        
        # ロックを解放する
        msvcrt.locking(lock_file.fileno(), msvcrt.LK_UNLCK, 1)
        # ロックファイルを閉じる
        lock_file.close()
        # ロックファイルを削除する
        os.remove(lock_file_path)
        
        sys.exit(0)
    
    current_date = datetime.now()
    
    # アクティブなシートを開く
    temp_sheet = workbook.create_sheet("temp")  
    day_int = current_date.month
    sheet = workbook[f"{day_int}月"]
    
    # 項目の作成
    temp_sheet['A1'] = '名前'
    temp_sheet['B1'] = 'ID'
    temp_sheet['C1'] = '出席状況'
    
    # エクセルにすべてのデータを入力
    for row in sheet.iter_rows(min_row=2, values_only=True):  # 2行目以降のデータを取得
        name = row[0]  # 名前
        display_id = row[1]  # 表示ID
        
        # 一時シートにデータを追加
        temp_sheet.append([name, display_id, None])  # 欠席状況の列は初めはNoneにしておく

    # 現在の日付を取得
    current_date_d = datetime.now().day  # 今日の日付（適宜変更）

    # コピー範囲の設定
    start_row = 2  # データの開始行
    end_row = len(sheet['A'])  # 最後の行まで
    start_column = current_date_d + 9  # 今日の日付に基づく開始列
    end_column = start_column  # 1列のみコピーする

    # コピー先の開始セルの指定
    dest_start_row = 2
    dest_start_column = 3  # 欠席状況の列（適宜変更）

    # 範囲をコピーしてコピー先のセルに貼り付ける
    for row in range(start_row, end_row + 1):
        for col in range(start_column, end_column + 1):
            cell_value = sheet.cell(row=row, column=col).value
            dest_row = row - start_row + dest_start_row
            dest_col = col - start_column + dest_start_column
            dest_cell = temp_sheet.cell(row=dest_row, column=dest_col)
            dest_cell.value = cell_value

    # 欠席と入力
    end_row = len(temp_sheet['A']) + 1  # データの数に基づいて終了行を決定する
    for row_number in range(2, end_row + 1):
        cell_value = temp_sheet.cell(row=row_number, column=3).value  # 欠席状況の列を取得
        if cell_value is None or cell_value == "":  # 空白またはNoneの場合
            temp_sheet.cell(row=row_number, column=3, value='無断欠席')  # 初めは無断欠席として設定

    # 保存
    ar_filename = f"{current_date_y}Attendance_records.xlsx"  # 保存するファイル名
    workbook.save(ar_filename)  # 変更を保存
    
    capbool = False
    
    def mainwindowshow(): #? メインウィンドウ表示
        
        # シートから欠席のデータを取得してリストに格納
        data = []
        
        for row in temp_sheet.iter_rows(values_only=True):
            if row[2] == '無断欠席':  # 無断欠席のデータのみを抽出 
                modified_row = list(row)
                modified_row[2] = '未出席'
                data.append(modified_row)
            else:
                data.append(list(row))
        
        header = ['名前', 'ID', '出席状況']  # 各列のヘッダーを指定
        
        # 左側のレイアウト
        left_column = [
            [sg.Text('記録なし', key="-OUTPUT-", font=("Helvetica", 40))],
            [sg.Text('名前またはIDを入力してください。', key="-INPUT-", font=("Helvetica", 15))],
            [sg.InputText(key="-NAME-", font=("Helvetica", 15))],
            [
                sg.Button('OK', bind_return_key=True, font=("Helvetica", 15)),
                sg.Button('終了', font=("Helvetica", 15)),
                sg.Checkbox('欠席', key='-ABSENCE-', enable_events=True),
                sg.Checkbox('早退', key='-LEAVE_EARLY-', enable_events=True)
            ],
            [sg.Table(key="-TABLE-", values=data, headings=header, display_row_numbers=False, auto_size_columns=False, num_rows=min(20, len(data)))],
            [sg.Text('', key="-INFO-", font=("Helvetica", 15))]
        ]

        # 右側のレイアウト（カメラ映像用）
        right_column = [
            [sg.Text("カメラ映像", font=("Helvetica", 20))],
            [sg.Image(filename="", key="-IMAGE-")]  # カメラ映像表示用
        ]

        # 全体レイアウト
        if config_data['facial_recognition']:
            layout = [
                [sg.Column(left_column), sg.VSeparator(), sg.Column(right_column)]
            ]
        else:
            layout = [
                [sg.Column(left_column)]
            ]
        
        window = sg.Window('出席処理', layout, finalize=True)
        window.Maximize()
        return window  # window変数を返す
    
    def get_name_by_id(id):  # IDから名前を取得
        # Excelファイルを開く
        wb = openpyxl.load_workbook('user_list.xlsx')
        sheet = wb['user_list']
        
        # IDが数値かどうかを確認
        if id.isdigit():  # 数値の場合はIDで検索
            for row in sheet.iter_rows(min_row=2, values_only=True):  # ヘッダーをスキップしてデータ行を処理
                try:
                    # IDが一致するかを比較
                    if float(row[0]) == float(id):  # ID列が数値の場合、floatに変換して比較
                        return row[1]  # 名前を返す
                except Exception as e:
                    logging.error(f"Error processing row {row}: {e}")
        else:  # 数値以外の場合は名前で検索
            for row in sheet.iter_rows(min_row=2, values_only=True):
                try:
                    # 名前が一致するかを比較
                    if str(row[1]).strip() == str(id).strip():  # 名前列が一致する場合
                        return row[1]  # 名前を返す
                except Exception as e:
                    logging.error(f"Error processing row {row}: {e}")
        
        return "記録されていない名前"  # 見つからない場合
    
    window = mainwindowshow()  # mainwindowshow()関数を呼び出して、window変数に代入する
    
    if config_data['facial_recognition']:
        # カメラを初期化
        video_capture = cv2.VideoCapture(0)

        process_this_frame = True
    face_permitted = False
    
    while True:  # 無限ループ
        # イベントとデータの読み込み
        event, values = window.read(timeout=20)
        
        if event == '-ABSENCE-':
            if values['-ABSENCE-']:
                window['-LEAVE_EARLY-'].update(False)
        
        if event == '-LEAVE_EARLY-':
            if values['-LEAVE_EARLY-']:
                window['-ABSENCE-'].update(False)
        
        if config_data['facial_recognition']:
            window["-INFO-"].update("顔認識は正常に稼働中です。")
            
            # 許容度（低いほど厳しい）
            tolerance = 0.5
            
            data_file = "face_data.pkl"  # 保存するデータファイル名
            
            # 既存データをロード
            face_data = load_face_data(data_file)
            known_face_encodings, known_face_names = face_data
            
            ret, frame = video_capture.read()
            if not ret:
                logging.error("カメラの映像取得に失敗しました")
                break
            
            # 顔認識処理
            if process_this_frame:
                small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
                rgb_small_frame = np.ascontiguousarray(small_frame[:, :, ::-1])
                face_locations = face_recognition.face_locations(rgb_small_frame)
                face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)
                face_locations = [(top * 4, right * 4, bottom * 4, left * 4) for (top, right, bottom, left) in face_locations]
            
            process_this_frame = not process_this_frame

            # 顔認識結果をフレームに描画
            for (top, right, bottom, left), face_encoding in zip(face_locations, face_encodings):
                matches = face_recognition.compare_faces(known_face_encodings, face_encoding, tolerance=tolerance)
                name = "Unknown"
                
                if True in matches:
                    first_match_index = matches.index(True)
                    name = known_face_names[first_match_index]
                    
                    face_permitted = True
                    break
                
                cv2.rectangle(frame, (left, top), (right, bottom), (0, 0, 255), 2)
                font = cv2.FONT_HERSHEY_DUPLEX
                cv2.putText(frame, name, (left + 6, bottom - 6), font, 0.5, (255, 255, 255), 1)
            
            # カメラ映像の更新
            imgbytes = cv2.imencode(".png", frame)[1].tobytes()
            window["-IMAGE-"].update(data=imgbytes)

        # OKボタン押されたときの処理
        if event == 'OK' or event == 'Escape:13' or capbool or face_permitted:
            
            if face_permitted == False:
                name = values["-NAME-"]
            face_permitted = False
            
            # 入力されたIDでExcelから検索
            name = get_name_by_id(name)
            
            if name == "記録されていない名前":
                # 名前が見つからない場合の処理
                messagebox.showwarning("WARNING", "IDに対応する名前が見つかりません。")
                continue
            
            # コンフィグの値を取得
            absence_state = values['-ABSENCE-']
            leave_early = values['-LEAVE_EARLY-']
            
            current_date = datetime.now()
            
            AttendanceTime = f"出席 {current_date.strftime('%H')}:{current_date.strftime('%M')}"
            info = "出席"
            
            for row in range(1, temp_sheet.max_row + 1):
                if temp_sheet.cell(row=row, column=1).value == name:
                    if temp_sheet.cell(row=row, column=3).value != "無断欠席":
                        if absence_state == False:
                            if leave_early == False:
                                info = "出席済み"
            
            # 状態を記録
            
            if absence_state:
                AttendanceTime = "欠席"
                info = "欠席"
            elif leave_early:
                AttendanceTime = f"早退 {current_date.strftime('%H')}:{current_date.strftime('%M')}"
                info = "早退"
            elif current_date.hour > lateness_time_hour or (current_date.hour == lateness_time_hour and current_date.minute > lateness_time_minute):
                AttendanceTime = f"遅刻 {current_date.strftime('%H')}:{current_date.strftime('%M')}"
                info = "遅刻"
            
            if info == "error":
                messagebox.showwarning('警告', '早退または欠席、一つを選択してください。')
            elif info == "出席済み":
                messagebox.showinfo('INFO', f'{name}さんは既に出席済みです。')
            elif info == "欠席" or info == "早退":
                if messagebox.askyesno('INFO', f'{info}として{name}さんを記録しますか？'):
                    # 名前が一致する行を探し、出席を記録
                    for row in range(1, sheet.max_row + 1):
                        if sheet.cell(row=row, column=1).value == name:
                            sheet.cell(row=row, column=current_date.day + 9, value=AttendanceTime)
                            workbook.save(ar_filename)
                    
                    for row in range(1, temp_sheet.max_row + 1):
                        if temp_sheet.cell(row=row, column=1).value == name:
                            temp_sheet.cell(row=row, column=3, value=AttendanceTime)
                            workbook.save(ar_filename)
                            
                            # シートから欠席のデータを取得してリストに格納
                            data = []
                            
                            for row in temp_sheet.iter_rows(values_only=True):
                                if row[2] == '無断欠席':  # 無断欠席のデータのみを抽出 
                                    modified_row = list(row)
                                    modified_row[2] = '未出席'
                                    data.append(modified_row)
                                else:
                                    data.append(list(row))
                            
                            window["-TABLE-"].update(data)
                            
                            window["-LEAVE_EARLY-"].update(False)
                            window["-ABSENCE-"].update(False)
                            
                            window["-OUTPUT-"].update(f'{name}さんの{info}処理は完了しました。')
                            
                            window["-NAME-"].update("")  # 入力フィールドをクリア
                            break
            else:
                # 名前が一致する行を探し、出席を記録
                for row in range(1, sheet.max_row + 1):
                    if sheet.cell(row=row, column=1).value == name:
                        sheet.cell(row=row, column=current_date.day + 9, value=AttendanceTime)
                        workbook.save(ar_filename)
                
                for row in range(1, temp_sheet.max_row + 1):
                    if temp_sheet.cell(row=row, column=1).value == name:
                        temp_sheet.cell(row=row, column=3, value=AttendanceTime)
                        workbook.save(ar_filename)
                        
                        # シートから欠席のデータを取得してリストに格納
                        data = []
                        
                        for row in temp_sheet.iter_rows(values_only=True):
                            if row[2] == '無断欠席':  # 無断欠席のデータのみを抽出 
                                modified_row = list(row)
                                modified_row[2] = '未出席'
                                data.append(modified_row)
                            else:
                                data.append(list(row))
                        
                        window["-TABLE-"].update(data)
                        window["-LEAVE_EARLY-"].update(False)
                        window["-ABSENCE-"].update(False)
                        window["-OUTPUT-"].update(f'{name}さんの{info}処理は完了しました。')
                        window["-NAME-"].update("")  # 入力フィールドをクリア
                        break
        
        # ウィンドウが閉じられた場合の処理
        if event == '終了' or event == sg.WIN_CLOSED:
            # カメラを解放
            if config_data['facial_recognition']:
                video_capture.release()
            
            window.close()
            
            # 一時シート削除
            sheet_name_to_delete = "temp"
            
            if sheet_name_to_delete in workbook.sheetnames:
                sheet_to_delete = workbook[sheet_name_to_delete]
                workbook.remove(sheet_to_delete)
            else:
                logging.warning("Temporary sheet deletion failure.")
            
            # 時間変数の設定
            current_date = datetime.now()
            current_date_y = current_date.strftime("%Y")
            
            # 記録ファイル名
            ar_filename = f"{current_date_y}Attendance_records.xlsx"
            
            # すべての行の3列目のセルが空白の場合「無断欠席」を記録
            for row in range(2, sheet.max_row + 1):
                if sheet.cell(row=row, column=current_date.day + 9).value == None:
                    sheet.cell(row=row, column=current_date.day + 9, value="無断欠席")
            
            # 変更を保存する
            workbook.save(ar_filename)
            
            messagebox.showinfo('完了', '記録終了は正常に終了しました。')
            break


def setting(): #? 設定変更画面
    
    while True:
        
        Automatic_late_time_setting = config_data['automatic_late_time_setting']
        Manual_late_time_setting = not Automatic_late_time_setting
        Lateness_Time = config_data['lateness_time']
        facial_recognition = config_data['facial_recognition']
        
        # レイアウトの定義
        layout = [
            [sg.Text('変更したい設定だけ変更してください。')],
            [sg.Text('遅刻時間')],
            [
                sg.Checkbox('起動時の時間 + X 分後に自動的に決める。', default=Automatic_late_time_setting, key='-automatic_late_time_setting-', enable_events=True),
                sg.Checkbox('時間を手動で入力する。', default=Manual_late_time_setting, key='-ManualLateTimeSetting-', enable_events=True)
            ],
            [sg.Text('自動設定の場合の X を決めてください: '), sg.InputText(default_text=Lateness_Time, key="-LatenessTime-", disabled=Manual_late_time_setting, disabled_readonly_background_color='grey', enable_events=True)], 
            [sg.Checkbox('顔認識', default=facial_recognition, key='-facial_recognition-', enable_events=True)],
            [sg.Text('顔認識は環境によっては正常に動かない場合があります。また処理が重くなる場合があります。')],
            [sg.Button('戻る'), sg.Button('初期設定に戻す')]
        ]
        
        # ウィンドウの生成
        window = sg.Window('設定', layout)
        
        # イベントループ
        while True:
            event, values = window.read()
            
            Lateness_Time = values['-LatenessTime-']
            
            if event == '初期設定に戻す':
                window.close()
                
                config_data['automatic_late_time_setting'] = True
                config_data['lateness_time'] = 15
                
                config_data["facial_recognition"] = False
                
                json_save()
                
                return
            
            if event == sg.WINDOW_CLOSED or event == '戻る': # 終了
                window.close()
                
                if not Lateness_Time == None:
                    lateness_time_str = Lateness_Time  # InputTextウィジェットからの文字列を取得
                    lateness_time_int = int(lateness_time_str)  # 文字列をint型に変換
                    
                    config_data['lateness_time'] = lateness_time_int
                
                json_save()
                
                return
            
            if event == '-facial_recognition-':
                if values['-facial_recognition-']:
                    # config変更
                    config_data["facial_recognition"] = True
                else:
                    # config変更
                    config_data["facial_recognition"] = False
            
            if event == '-automatic_late_time_setting-':
                if values['-automatic_late_time_setting-']:
                    window['-ManualLateTimeSetting-'].update(False)
                    
                    # 入力ボックス有効化
                    window['-LatenessTime-'].update(disabled=False)
                    
                    # config変更
                    config_data["automatic_late_time_setting"] = True
            
            elif event == '-ManualLateTimeSetting-':
                if values['-ManualLateTimeSetting-']:
                    window['-automatic_late_time_setting-'].update(False)
                    
                    # 入力ボックス無効化
                    window['-LatenessTime-'].update(disabled=True)
                    
                    # config変更
                    config_data["automatic_late_time_setting"] = False

#? 起動

window['-PROG-'].update(30)

#? 初期設定

# Tkinterウィンドウを作成
root = tk.Tk()
# topmost指定(最前面)
root.attributes('-topmost', True)
root.withdraw()
root.lift()
root.focus_force()

window['-PROG-'].update(40)

#? ファイルの存在確認

# configファイル
if not os.path.exists("config.json"):
    exit_with_error("config.json file not found.")

if not os.path.exists("user_list.xlsx"):
    exit_with_error("user_list.xlsx file not found.")

window['-PROG-'].update(50)

# 記録ファイル
current_date = datetime.now()
current_date_y = current_date.strftime("%Y")

ar_filename = f"{current_date_y}Attendance_records.xlsx"

if not os.path.exists(ar_filename):
    # ファイルが存在しない場合の処理
    record_file_creation()

window['-PROG-'].update(60)

#? config読み込み

# 設定ファイルのパス
config_file_path = 'config.json'

# 設定ファイルの読み込み
with open(config_file_path, 'r') as config_file:
    config_data = json.load(config_file)


window['-PROG-'].update(70)

#? 顔認証

if config_data['facial_recognition']:
    
    data_file = "face_data.pkl"  # 保存するデータファイル名
    
    face_permit = False
    
    # 顔データをロードする
    def load_face_data(file):
        if os.path.exists(file):
            with open(file, "rb") as f:
                return pickle.load(f)
        return None
    
    # 顔データを保存する
    def save_face_data(file, encodings, names):
        with open(file, "wb") as f:
            pickle.dump((encodings, names), f)
            
    # 既存データをロード
    face_data = load_face_data(data_file)
    
    window['-PROG-'].update(80)
    
    if face_data:
        known_face_encodings, known_face_names = face_data
    else:
        logging.info("顔データが見つからないため、新たに作成します。")
        # 画像フォルダのパス
        image_folder = 'face_images/'
        
        # 画像を読み込み、顔の特徴を記録する
        known_face_encodings = []
        known_face_names = []
        
        # フォルダごとに画像を処理
        folder_list = [folder for folder in os.listdir(image_folder) if os.path.isdir(os.path.join(image_folder, folder))]
        
        window['-PROG-'].update(90)
        
        # フォルダの処理
        for foldername in folder_list:
            folder_path = os.path.join(image_folder, foldername)
            
            # フォルダ内の画像ファイルを処理
            image_files = [f for f in os.listdir(folder_path) if f.endswith(".jpg")]
            
            # フォルダ内の画像ファイルの処理
            for filename in image_files:
                image_path = os.path.join(folder_path, filename)
                image = face_recognition.load_image_file(image_path)
                encoding = face_recognition.face_encodings(image)
                
                if encoding:
                    # 複数の画像から同じ人物の特徴を追加
                    known_face_encodings.append(encoding[0])
                    known_face_names.append(foldername)  # フォルダ名を表示名として取得
        
        # データを保存
        save_face_data(data_file, known_face_encodings, known_face_names)

window['-PROG-'].update(100)

window.close()

#? メイン

update()

while True:  #? 無限ループ
    # GUI画面のレイアウト
    layout = [
        [sg.Text("起動する機能を選んでください。", font=("Helvetica", 15), justification='center')],  # カンマを追加
        [sg.Button('記録', bind_return_key=True, font=("Helvetica", 15)),
            sg.Button('設定', bind_return_key=True, font=("Helvetica", 15)),
            sg.Button('終了', bind_return_key=True, font=("Helvetica", 15))]
    ]
    
    menu = sg.Window('MENU', layout, finalize=True, keep_on_top=True)
    
    event, values = menu.read()
    
    if event == sg.WIN_CLOSED or event == '終了':  # Xボタンが押されたか、'終了'ボタンが押された場合
        menu.close()
        # JSONデータを保存
        with open('config.json', 'w') as f:
            json.dump(config_data, f, indent=4)
        
        #? 重複起動関係
        
        # ロックを解放する
        msvcrt.locking(lock_file.fileno(), msvcrt.LK_UNLCK, 1)
        # ロックファイルを閉じる
        lock_file.close()
        # ロックファイルを削除する
        os.remove(lock_file_path)
        
        sys.exit(0)
    
    if event == '記録':
        menu.close()

        record()
    
    if event == '設定':
        menu.close()
        tk.Tk().withdraw()
        setting()
