#? インポート
import PySimpleGUI as sg
import sys
import os
from datetime import datetime
from tkinter import messagebox
import sqlite3
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
import cv2
import time

#? エラー時の処理の作成
def exit_with_error(message):
    print(f"Error: {message}")
    messagebox.showerror("Error:", message)
    sys.exit(1)  # アプリケーションをエラーコード 1 で終了します

#? 変数の初期設定

name_column_index = None
date_row_index = None

# 時間変数の設定
current_date = datetime.now()
current_date_y = current_date.strftime("%Y")
current_date_m = current_date.strftime("%m")
current_date_d = current_date.strftime("%d")
current_date_h = current_date.strftime("%H")
current_date_M = current_date.strftime("%M")
current_date_s = current_date.strftime("%S")

# 一時ファイル名
ar_filename = f"{current_date_y}Attendance records.xlsx"

#? Excel初期設定

# 新しいWorkbook（エクセルファイル）を作成して、ファイル名を指定
try:
    workbook = load_workbook(ar_filename)
except FileNotFoundError:
    exit_with_error("File not found")

# アクティブなシートを開く
temp_sheet = workbook.create_sheet("temp")  
day_int = int(current_date_m)
sheet = workbook[f"{day_int}月"]

# 項目の作成
temp_sheet['A1'] = '名前'
temp_sheet['B1'] = '学年'
temp_sheet['C1'] = '出席状況'

#? DB初期設定

# DBに接続する

conn = sqlite3.connect('Register.db')
cursor = conn.cursor()    


# すべてのデータを取得
cursor.execute('SELECT Name, GradeinSchool FROM Register')
all_data = cursor.fetchall()

#? 一時ファイルにDBを入力

# エクセルにすべてのデータを入力
for data in all_data:
    row_number = temp_sheet.max_row + 1
    temp_sheet.cell(row=row_number, column=1, value=data[0])  # 名前
    temp_sheet.cell(row=row_number, column=2, value=data[1])  # 学年

end_row = len(all_data) # データの数に基づいて終了行を決定する
for row_number in range(2, end_row + 2): # 2からend_row + 1までの行数を繰り返し処理
    temp_sheet.cell(row=row_number, column=3, value='未出席') # 初めは未出席として設定


# 保存
workbook.save(ar_filename)

information = '記録なし'

#? 二次元コード
# カメラを起動
cap = cv2.VideoCapture(1)

# QRコードを検出するためのdetectorを作成
detector = cv2.QRCodeDetector()

# last_qr_dataの初期化
last_qr_data = None

capbool = False

def mainwindowshow(): #? メインウィンドウ表示
    
    # シートから未出席のデータを取得してリストに格納
    data = []
    
    
    
    for row in temp_sheet.iter_rows(values_only=True):
        if row[2] == '未出席':  # 未出席のデータのみを抽出
            data.append(list(row))  # タプルからリストに変換して追加
            
    # ヘッダーを取得
    header = list(temp_sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    
    left_column = [
        [sg.Text(information, font=("Helvetica", 40))],
        [sg.Text('苗字を入力してください。', key="-INPUT-", font=("Helvetica", 15))],
        [sg.InputText(key="-NAME-", font=("Helvetica", 15))],
        [sg.Button('OK', bind_return_key=True, font=("Helvetica", 15)), sg.Button('終了', bind_return_key=True, font=("Helvetica", 15))],
        [sg.Table(values=data, headings=header, display_row_numbers=False, auto_size_columns=False, num_rows=min(20, len(data)))]
    ]
    
    right_column = [
        [sg.Image(filename='', key="-IMAGE-")],
        [sg.Text('', key="-QR_DATA-")],
    ]
    
    # GUI画面のレイアウト
    layout = [
        [sg.Column(left_column), sg.Column(right_column)]
    ]
    
    window = sg.Window('出席処理', layout, finalize=True, disable_close=True)
    window.Maximize()
    return window  # window変数を返す

def get_name_by_id(input_id): #? IDから名前を取得
    # IDに対応するnameをクエリで検索
    cursor.execute("SELECT Name FROM Register WHERE ID=?", (input_id,))
    result = cursor.fetchone()  # 一致する最初の行を取得
    if result:
        return result[0]  # nameを返す
    else:
        return "IDに対応する名前が見つかりません"

window = mainwindowshow()  # mainwindowshow()関数を呼び出して、window変数に格納する

while True: #? 無限ループ
    # イベントとデータの読み込み
    event, values = window.read(timeout=20)
    
    # カメラからフレームを取得
    ret, frame = cap.read()
    
    # QRコードを検出
    try:
        qr_data, _, _ = detector.detectAndDecode(frame)
    except cv2.error as e:
        print("QRコードの検出中にエラーが発生しました:", e)
        qr_data = None
    
    # 読み取ったQRコードがあれば
    if qr_data and qr_data != last_qr_data:
        window['-QR_DATA-'].update(f'QRコードの中の数値: {qr_data}')
        last_qr_data = qr_data
        capbool = True
        name = qr_data
        print(f"QRコードの中の数値: {qr_data}")
    
    # OpenCVのBGR形式をPySimpleGUIの画像形式に変換してウィンドウに表示
    if ret:
        resized_frame = cv2.resize(frame, (400, 300))
        imgbytes = cv2.imencode('.png', resized_frame)[1].tobytes()
        window['-IMAGE-'].update(data=imgbytes)
    
    #? ウィンドウを閉じる時の処理
    if event == sg.WIN_CLOSED:
        messagebox.showinfo('警告', 'windowを閉じるのは「終了」ボタンから行ってください。')
        pass
    
    #? OKが押されたときの処理
    if event == 'OK' or event == 'Escape:13' or capbool:
        capbool = False
        if not name:
            name = values["-NAME-"]
        
        if name.isdigit():
            name = get_name_by_id(int(name))
            result = name
        else:
            cursor.execute('SELECT GradeinSchool FROM Register WHERE Name = ?', (name,))
            result = cursor.fetchone()
        
        
        
        print('入力された値：', name)
        if result:
            # 名前が一致する行を探し、出席を記録
            for row in range(1, temp_sheet.max_row + 1):
                if temp_sheet.cell(row=row, column=1).value == name:
                    temp_sheet.cell(row=row, column=3, value='出席')
                    workbook.save(ar_filename)
                    
                    information = f'{name}さんの出席処理は完了しました。'
                    window["-NAME-"].update("")  # 入力フィールドをクリア
                    conn.commit()  # 変更を確定
                    
                    # ウィンドウを閉じてから新しいウィンドウを作成
                    window.close()
                    window = mainwindowshow()
                    break
        else:
            # 失敗処理
            print(f"{name} はデータベースに存在しません。")
            messagebox.showinfo('失敗', f'{name} はデータベースに存在しません。')
            window["-NAME-"].update("")  # 入力フィールドをクリア
    
    #? 閉じられるときの処理
    if event == '終了':
        
        # カメラを解放
        cap.release()
        
        window.close()
        
        # 警告メッセージ表示
        endresult = messagebox.askquestion('警告', '本当に閉じますか？', icon='warning')
        if endresult == 'yes': # yes
            
            start_row = 2
            end_row = 26
            start_column = 3
            end_column = 3
            
            # コピー先の開始セルの指定
            dest_start_row = 2
            dest_start_column = int(current_date_d) + 2 # 文字列を整数値に変換
            
            # 範囲をコピーしてコピー先のセルに貼り付ける
            for row in range(start_row, end_row + 1):
                for col in range(start_column, end_column + 1):
                    cell_value = temp_sheet.cell(row=row, column=col).value
                    dest_row = row - start_row + dest_start_row
                    dest_col = col - start_column + dest_start_column
                    dest_cell = sheet.cell(row=dest_row, column=dest_col)
                    # セルの値をコピー
                    dest_cell.value = cell_value
            
            
            workbook.save(ar_filename)
            
            #? 一時シート削除
            
            # 削除したいシート名を指定
            sheet_name_to_delete = "temp"
            
            # シートを削除
            if sheet_name_to_delete in workbook.sheetnames:
                sheet_to_delete = workbook[sheet_name_to_delete]
                workbook.remove(sheet_to_delete)
                print(f"{sheet_name_to_delete} シートを削除しました。")
            else:
                print(f"{sheet_name_to_delete} シートは存在しません。")
            
            workbook.save(ar_filename)
            
            messagebox.showinfo('完了', '記録終了は正常に終了しました。')
            conn.close()
            break
        else:
            # ウィンドウを閉じてから新しいウィンドウを作成
            window.close()
            window = mainwindowshow()
            continue