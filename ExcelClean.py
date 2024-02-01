#? インポート
import PySimpleGUI as sg
import sys
import os
from datetime import datetime
from tkinter import messagebox
import sqlite3
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
import cv2
import time
import json
import subprocess

print("Color Change >>> ", end="")

#? エラー時の処理の作成

def exit_with_error(message):
    print(f"Error: {message}")
    messagebox.showerror("Error:", message)
    sys.exit(1)  # アプリケーションをエラーコード 1 で終了します

# 時間変数の設定
current_date = datetime.now()
current_date_y = current_date.strftime("%Y")

# 一時ファイル名
ar_filename = f"{current_date_y}Attendance records.xlsx"

try:
    workbook = load_workbook(ar_filename)
except FileNotFoundError:
    exit_with_error("File not found")

# 出席と欠席のセルの背景色を定義
absent_fill = PatternFill(start_color='FFC0CB', end_color='FFC0CB', fill_type='solid')  # 赤色
present_fill = PatternFill(start_color='ADFF2F', end_color='ADFF2F', fill_type='solid')  # 緑色

for sheet in workbook.sheetnames:
    current_sheet = workbook[sheet]
    
    # すべてのセルを調べる
    for row in current_sheet.iter_rows():
        for cell in row:
            # セルの値が欠席か出席かを確認し、背景色を変更する
            if cell.value == '欠席':
                cell.fill = absent_fill
            elif cell.value == '出席':
                cell.fill = present_fill

# 変更を保存する
workbook.save(ar_filename)

print("done")

messagebox.showinfo('完了', '正常に終了しました。')

subprocess.run(["python", "Menu.py"])
sys.exit(0)