# インポート
import PySimpleGUI as sg

import os
from datetime import datetime
from tkinter import messagebox
import sqlite3
import openpyxl
from openpyxl.styles import PatternFill

# 新しいWorkbook（エクセルファイル）を作成して、ファイル名を指定
workbook = openpyxl.Workbook()
# アクティブなシートを開く
sheet = workbook.active

# 項目の作成
sheet['A1'] = '名前'
sheet['B1'] = '学年'
sheet['C1'] = '出席状況'

# DBに接続する
conn = sqlite3.connect('Register.db')
cursor = conn.cursor()

# すべてのデータを取得
cursor.execute('SELECT Name, GradeinSchool FROM Register')
all_data = cursor.fetchall()

# エクセルにすべてのデータを入力
for data in all_data:
    row_number = sheet.max_row + 1
    sheet.cell(row=row_number, column=1, value=data[0])  # 名前
    sheet.cell(row=row_number, column=2, value=data[1])  # 学年
    sheet.cell(row=row_number, column=3, value='未出席')  # 初めは未出席として設定

# 保存
workbook.save('temp.xlsx')

information = '記録なし'

def mainwindowshow():
    # シートから未出席のデータを取得してリストに格納
    data = []
    
    for row in sheet.iter_rows(values_only=True):
        if row[2] == '未出席':  # 未出席のデータのみを抽出
            data.append(list(row))  # タプルからリストに変換して追加
            
    # ヘッダーを取得
    header = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    
    # GUI画面のレイアウト
    layout = [
        [sg.Text(information, font=("Helvetica", 15))],
        [sg.Text('苗字を入力してください。', key="-INPUT-", font=("Helvetica", 15))],
        [sg.InputText(key="-NAME-", font=("Helvetica", 15))],
        [sg.Button('OK', bind_return_key=True, font=("Helvetica", 15))],
        [sg.Table(values=data, headings=header, display_row_numbers=False, auto_size_columns=False, num_rows=min(20, len(data)))]
    ]
    
    window = sg.Window('出席処理', layout, finalize=True)
    window.Maximize()
    return window  # window変数を返す

window = mainwindowshow()  # mainwindowshow()関数を呼び出して、window変数に格納する

while True:
    # ウィンドウの入力値を読み取る
    event, values = window.read()
    
    if event == sg.WIN_CLOSED: # 閉じられるときの処理
        
        # 警告メッセージ表示
        endresult = messagebox.askquestion('警告', '本当に閉じますか？', icon='warning')
        if endresult == 'yes':
            # 行ごとに条件を確認し、条件が満たされた場合に背景色を変更-
            for row_number in range(2, sheet.max_row + 1):
                attendance_status = sheet.cell(row=row_number, column=3).value
                
                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row_number, column=col)
                    
                    if attendance_status == '出席':
                        cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # 背景色を緑に設定
                    else:
                        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 背景色を黄色に設定
            
            # 変更を保存
            workbook.save('temp.xlsx')
            
            # 現在の日付を取得
            current_date = datetime.now().strftime("%Y-%m-%d")
            
            # 変更前のファイル名
            old_filename = "temp.xlsx"
            
            # 新しいファイル名を作成
            new_filename = f"{current_date}.xlsx"
            
            # ファイル名の変更
            os.rename(old_filename, new_filename)
            
            print(f"ファイル名を変更しました: {old_filename} → {new_filename}")
            messagebox.showinfo('完了', '記録終了は正常に終了しました。')
            
            conn.close()
            break
        else:
            # ウィンドウを閉じてから新しいウィンドウを作成
            window.close()
            window = mainwindowshow()
            continue
    
    if event == 'OK' or event == 'Escape:13': # OKが押されたときの処理
        name = values["-NAME-"]
        cursor.execute('SELECT GradeinSchool FROM Register WHERE Name = ?', (name,))
        result = cursor.fetchone()
        
        print('入力された値：', name)
        
        if result:
            print(f"{name} はデータベースに存在します。")
            
            # 名前が一致する行を探し、出席を記録
            for row in range(1, sheet.max_row + 1):
                if sheet.cell(row=row, column=1).value == name:
                    sheet.cell(row=row, column=3, value='出席')
                    workbook.save('temp.xlsx')
                    information = f'{name}さんの出席処理は完了しました。'
                    window["-NAME-"].update("")  # 入力フィールドをクリア
                    conn.commit()  # 変更を確定
                    
                    # ウィンドウを閉じてから新しいウィンドウを作成
                    window.close()
                    window = mainwindowshow()
                    break
        else:
            # 失敗処理
            print(f"{name} はデータベースに存在しません。")
            messagebox.showinfo('失敗', f'{name} はデータベースに存在しません。')
            window["-NAME-"].update("")  # 入力フィールドをクリア