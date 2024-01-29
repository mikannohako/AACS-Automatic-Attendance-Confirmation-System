# インポート
import PySimpleGUI as sg
import os
from datetime import datetime
from tkinter import messagebox
import sqlite3
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

# 月ごとのシートを作成する関数
def create_month_sheet(workbook, month):
    sheet_name = month
    if month.startswith("0"):
        sheet_name = month[1:]
    if sheet_name not in workbook.sheetnames:
        workbook.create_sheet(sheet_name)
        # 作成したシートを返す
        return workbook[sheet_name]

# 日付データを入力する関数
def input_date_data(sheet, month):
    days_in_month = 30 if month in ["04", "06", "09", "11"] else 31 if month != "02" else 29  # 月ごとの日数
    for day in range(1, days_in_month + 1):
        sheet.cell(row=1, column=day + 2).value = day
    # 学年の列を追加
    sheet.cell(row=1, column=1).value = '名前'
    sheet.cell(row=1, column=2).value = '学年'

# 新しいWorkbook（エクセルファイル）を作成して、ファイル名を指定
workbook = openpyxl.Workbook()

# DBに接続する
conn = sqlite3.connect('Register.db')
cursor = conn.cursor()

# すべてのデータを取得
cursor.execute('SELECT Name, GradeinSchool FROM Register')
all_data = cursor.fetchall()

# 月ごとのシートを作成して名前と学年の行を追加
for month in range(1, 13):
    month_str = str(month).zfill(2)  # 1桁の月を2桁の文字列に変換
    sheet = create_month_sheet(workbook, f"{month_str}月")
    input_date_data(sheet, month_str)
    
    # エクセルにすべてのデータを入力
    for data in all_data:
        row_number = sheet.max_row + 1
        sheet.cell(row=row_number, column=1, value=data[0])  # 名前
        sheet.cell(row=row_number, column=2, value=data[1])  # 学年
    
    table = openpyxl.worksheet.table.Table(displayName=f"Table{month}", ref="A1:B26")
    
    # スタイル設定
    style = TableStyleInfo(
        name="TableStyleMedium9",  # テーブルのスタイル名
        showRowStripes=True,    # 行のストライプを表示する
    )
    
    # テーブルにスタイルを適用
    table.tableStyleInfo = style
    
    # シートにテーブルを追加
    sheet.add_table(table)

# 削除したいシート名を指定
sheet_name_to_delete = "Sheet"

# シートを削除
if sheet_name_to_delete in workbook.sheetnames:
    sheet_to_delete = workbook[sheet_name_to_delete]
    workbook.remove(sheet_to_delete)
    print(f"{sheet_name_to_delete} シートを削除しました。")
else:
    print(f"{sheet_name_to_delete} シートは存在しません。")

current_date = datetime.now().strftime("%Y")
# エクセルファイルを保存
workbook.save(f"{current_date}Attendance records.xlsx")