import os
from datetime import datetime
from tkinter import messagebox
import openpyxl
from openpyxl.styles import PatternFill

# 既存のエクセルファイルを開く
workbook = openpyxl.load_workbook('temp.xlsx')

# ワークブック内の最初のシートを取得
sheet = workbook.active

# 行ごとに条件を確認し、条件が満たされた場合に背景色を変更
for row_number in range(2, sheet.max_row + 1):
    attendance_status = sheet.cell(row=row_number, column=4).value
    
    for col in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=row_number, column=col)
        
        if attendance_status == '出席':
            cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # 背景色を緑に設定
        else:
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 背景色を黄色に設定

# 変更を保存
workbook.save('temp.xlsx')

# 現在の日付を取得
current_date = datetime.now().strftime("%Y-%m-%d")

# 変更前のファイル名
old_filename = "temp.xlsx"

# 新しいファイル名を作成
new_filename = f"{current_date}.xlsx"

# ファイル名の変更
os.rename(old_filename, new_filename)

print(f"ファイル名を変更しました: {old_filename} → {new_filename}")
messagebox.showinfo('完了', '記録終了は正常に終了しました。')